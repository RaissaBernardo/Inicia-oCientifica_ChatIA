"""
avaliar_rag.py - Avalia o desempenho do RAG medindo a métrica de Faithfulness.

COMO USAR:
----------
1. Preencha o arquivo golden_dataset.csv com suas perguntas e respostas do manual.
2. Certifique-se que o servidor está rodando:
       python -m uvicorn main:app --reload
3. Em outro terminal, rode este script:
       python avaliar_rag.py

O script vai:
  - Para cada caso (A a E), re-gerar o banco vetorial com novos parâmetros
  - Perguntar ao chatbot cada questão do golden dataset
  - Calcular o Faithfulness de cada resposta
  - Salvar TUDO numa planilha Excel: resultados_avaliacao.xlsx

ATENÇÃO: O servidor (uvicorn) precisa estar rodando em paralelo!
"""

import csv
import json
import subprocess
import sys
import time
from pathlib import Path

import httpx

# ── Tente importar openpyxl; instale se não tiver ────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[INFO] Instalando openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURAÇÕES — ajuste se necessário
# ─────────────────────────────────────────────────────────────────────────────
CHAT_URL         = "http://localhost:8000/chat"
GOLDEN_DATASET   = Path(__file__).parent / "golden_dataset.csv"
OUTPUT_XLSX      = Path(__file__).parent / "resultados_avaliacao.xlsx"
BACKEND_DIR      = Path(__file__).parent
INGEST_SCRIPT    = BACKEND_DIR / "ingest.py"

# Os 5 casos que o professor pediu
CASOS = [
    {"nome": "A", "chunk_size": 300,  "chunk_overlap": 30},
    {"nome": "B", "chunk_size": 900,  "chunk_overlap": 90},
    {"nome": "C", "chunk_size": 1200, "chunk_overlap": 120},
    {"nome": "D", "chunk_size": 1500, "chunk_overlap": 150},
    {"nome": "E", "chunk_size": 1800, "chunk_overlap": 180},
]

# ─────────────────────────────────────────────────────────────────────────────


def carregar_golden_dataset() -> list[dict]:
    """Lê o golden_dataset.csv e retorna lista de {id, pergunta, ground_truth}."""
    if not GOLDEN_DATASET.exists():
        print(f"\n[ERRO] Arquivo não encontrado: {GOLDEN_DATASET}")
        print("       Crie o golden_dataset.csv com as colunas: id,pergunta,ground_truth")
        sys.exit(1)

    dataset = []
    with open(GOLDEN_DATASET, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pergunta = row.get("pergunta", "").strip()
            ground_truth = row.get("ground_truth", "").strip()
            if pergunta and ground_truth:
                dataset.append({
                    "id": row.get("id", ""),
                    "pergunta": pergunta,
                    "ground_truth": ground_truth,
                })

    print(f"[✓] Golden Dataset carregado: {len(dataset)} perguntas.")
    return dataset


def perguntar_ao_chat(pergunta: str) -> tuple[str, list]:
    """Envia a pergunta ao servidor e retorna (resposta, fontes)."""
    try:
        resp = httpx.post(CHAT_URL, json={"question": pergunta}, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        return data.get("answer", ""), data.get("sources", [])
    except httpx.ConnectError:
        print("\n[ERRO] Não consegui conectar ao servidor.")
        print("       Certifique-se que o uvicorn está rodando:")
        print("       python -m uvicorn main:app --reload\n")
        sys.exit(1)
    except Exception as e:
        print(f"[AVISO] Erro ao perguntar: {e}")
        return "", []


def dividir_em_afirmacoes(texto: str) -> list[str]:
    """
    Quebra o texto em afirmações individuais.
    Estratégia simples: divide por '.', '!' e '?' e remove fragmentos curtos.
    """
    import re
    partes = re.split(r"(?<=[.!?])\s+", texto.strip())
    afirmacoes = [p.strip() for p in partes if len(p.strip()) > 15]
    return afirmacoes if afirmacoes else [texto.strip()]


def calcular_faithfulness(resposta: str, ground_truth: str) -> tuple[float, list[dict]]:
    """
    Calcula a métrica de Faithfulness:
        faithfulness = nº afirmações apoiadas / nº total de afirmações

    Para cada afirmação da resposta do chat, verifica se termos-chave
    da afirmação aparecem no ground truth (busca por palavras significativas).

    Retorna (score_0_a_1, lista_de_detalhes).
    """
    import re

    afirmacoes = dividir_em_afirmacoes(resposta)
    detalhes = []

    STOPWORDS = {
        "a", "o", "e", "de", "do", "da", "em", "um", "uma", "para", "com",
        "que", "se", "no", "na", "os", "as", "é", "ao", "dos", "das", "por",
        "seu", "sua", "ser", "não", "ou", "mais", "este", "esta", "isso",
        "deve", "deve", "ser", "foi", "são", "também", "pode", "quando",
    }

    gt_lower = ground_truth.lower()

    for afirmacao in afirmacoes:
        # Extrai palavras significativas (> 3 letras, sem stopwords)
        palavras = re.findall(r'\b[a-záéíóúãõâêôçàü]{4,}\b', afirmacao.lower())
        palavras_sig = [p for p in palavras if p not in STOPWORDS]

        if not palavras_sig:
            # Afirmação sem palavras significativas → considerada não apoiada
            detalhes.append({"afirmacao": afirmacao, "apoiada": False, "motivo": "sem palavras-chave"})
            continue

        # Conta quantas palavras significativas aparecem no ground truth
        encontradas = sum(1 for p in palavras_sig if p in gt_lower)
        taxa = encontradas / len(palavras_sig)

        # Critério: ≥ 50% das palavras-chave devem estar no ground truth
        apoiada = taxa >= 0.5
        detalhes.append({
            "afirmacao": afirmacao,
            "apoiada": apoiada,
            "motivo": f"{encontradas}/{len(palavras_sig)} palavras-chave encontradas ({taxa:.0%})"
        })

    apoiadas = sum(1 for d in detalhes if d["apoiada"])
    score = apoiadas / len(detalhes) if detalhes else 0.0
    return score, detalhes


def regerar_banco(chunk_size: int, chunk_overlap: int, nome_caso: str):
    """
    Gera o banco vetorial numa pasta separada por caso
    e pede ao servidor para trocar via /reload_db.
    Assim o servidor continua rodando normalmente.
    """
    pasta = f"chroma_db_caso_{nome_caso}"
    print(f"\n  ⚙️  Gerando banco na pasta '{pasta}'...")
    print("      Isso pode levar alguns minutos...")
    result = subprocess.run(
        [sys.executable, str(INGEST_SCRIPT),
         "--chunk_size",    str(chunk_size),
         "--chunk_overlap", str(chunk_overlap),
         "--output",        pasta],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"[ERRO] ingest.py falhou:\n{result.stderr}")
        sys.exit(1)
    print("  ✅ Banco gerado com sucesso!")

    print("  🔄 Avisando o servidor para usar o novo banco...")
    try:
        r = httpx.post(
            CHAT_URL.replace("/chat", "/reload_db"),
            json={"chroma_dir": pasta},
            timeout=30
        )
        r.raise_for_status()
        print("  ✅ Servidor trocou de banco!")
    except Exception as e:
        print(f"  ⚠️  /reload_db nao respondeu ({e}). Aguardando 10s...")
        time.sleep(10)


def avaliar_caso(caso: dict, dataset: list[dict]) -> list[dict]:
    """Roda todos os testes do golden dataset para um caso e retorna os resultados."""
    print(f"\n{'='*60}")
    print(f"  CASO {caso['nome']}  |  CHUNK_SIZE={caso['chunk_size']}  CHUNK_OVERLAP={caso['chunk_overlap']}")
    print(f"{'='*60}")

    regerar_banco(caso["chunk_size"], caso["chunk_overlap"], caso["nome"])

    resultados = []
    total = len(dataset)

    for i, item in enumerate(dataset, 1):
        print(f"  [{i:03d}/{total}] Perguntando: {item['pergunta'][:60]}...")
        resposta, _ = perguntar_ao_chat(item["pergunta"])

        if not resposta:
            score = 0.0
            detalhes = []
        else:
            score, detalhes = calcular_faithfulness(resposta, item["ground_truth"])

        print(f"         Faithfulness: {score:.1%}")

        resultados.append({
            "caso": caso["nome"],
            "chunk_size": caso["chunk_size"],
            "chunk_overlap": caso["chunk_overlap"],
            "id": item["id"],
            "pergunta": item["pergunta"],
            "ground_truth": item["ground_truth"],
            "resposta_chat": resposta,
            "faithfulness": score,
            "n_afirmacoes": len(detalhes),
            "n_apoiadas": sum(1 for d in detalhes if d["apoiada"]),
            "detalhe": "; ".join(
                f"[{'✓' if d['apoiada'] else '✗'}] {d['afirmacao'][:60]}... ({d['motivo']})"
                for d in detalhes
            ),
        })

    media = sum(r["faithfulness"] for r in resultados) / len(resultados) if resultados else 0
    print(f"\n  📊 Faithfulness médio do Caso {caso['nome']}: {media:.1%}")
    return resultados


# ─── Estilos da planilha ──────────────────────────────────────────────────────

COR_HEADER    = "1F4E79"   # azul escuro
COR_CASO      = {"A": "D6E4F7", "B": "D6F7E4", "C": "FFF3CD", "D": "FFE0CC", "E": "F0D6F7"}
COR_RESUMO    = "FCE4D6"
VERDE         = "70AD47"
VERMELHO      = "FF4B4B"
AMARELO       = "FFD966"

def cor_faithfulness(valor: float) -> str:
    if valor >= 0.7:
        return VERDE
    if valor >= 0.4:
        return AMARELO
    return VERMELHO

def estilo_celula(ws, cell_ref, valor=None, bold=False, bg=None, wrap=False, alinhamento="left"):
    cell = ws[cell_ref] if isinstance(cell_ref, str) else cell_ref
    if valor is not None:
        cell.value = valor
    if bold:
        cell.font = Font(bold=True, color="FFFFFF" if bg and bg == COR_HEADER else "000000")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical="top",
                                horizontal=alinhamento)

def borda_fina():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def salvar_xlsx(todos_resultados: list[dict]):
    wb = openpyxl.Workbook()

    # ── Aba RESUMO ────────────────────────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    # Título
    ws_resumo.merge_cells("A1:F1")
    c = ws_resumo["A1"]
    c.value = "Avaliação de Desempenho do RAG — Faithfulness por Caso"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 30

    # Cabeçalho da tabela de resumo
    headers = ["Caso", "CHUNK_SIZE", "CHUNK_OVERLAP", "Nº Perguntas", "Faithfulness Médio", "Avaliação"]
    for col, h in enumerate(headers, 1):
        cell = ws_resumo.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")
        cell.border = borda_fina()

    # Agrupa por caso e calcula média
    from collections import defaultdict
    por_caso = defaultdict(list)
    for r in todos_resultados:
        por_caso[r["caso"]].append(r)

    melhor_caso = None
    melhor_score = -1

    for linha, caso in enumerate(CASOS, 4):
        nome = caso["nome"]
        regs = por_caso[nome]
        media = sum(r["faithfulness"] for r in regs) / len(regs) if regs else 0

        if media > melhor_score:
            melhor_score = media
            melhor_caso = nome

        avaliacao = "Ótimo 🏆" if media >= 0.7 else ("Regular ⚠️" if media >= 0.4 else "Ruim ❌")

        valores = [nome, caso["chunk_size"], caso["chunk_overlap"], len(regs),
                   f"{media:.1%}", avaliacao]
        for col, v in enumerate(valores, 1):
            cell = ws_resumo.cell(row=linha, column=col, value=v)
            cell.fill = PatternFill("solid", fgColor=COR_CASO.get(nome, "FFFFFF"))
            cell.border = borda_fina()
            if col == 5:
                cell.font = Font(bold=True, color=cor_faithfulness(media))
            cell.alignment = Alignment(horizontal="center")

    # Destaque do melhor caso
    if melhor_caso:
        ws_resumo.merge_cells("A10:F10")
        c = ws_resumo["A10"]
        c.value = f"🏆 Melhor caso: {melhor_caso}  (Faithfulness médio: {melhor_score:.1%})"
        c.font = Font(bold=True, size=12)
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        c.alignment = Alignment(horizontal="center")

    # Larguras
    for col, w in zip("ABCDEF", [8, 14, 16, 16, 22, 18]):
        ws_resumo.column_dimensions[get_column_letter(col.index("A") + "ABCDEF".index(col) + 1)].width = w

    # ── Uma aba por caso ──────────────────────────────────────────────────────
    cab_det = ["ID", "Pergunta", "Ground Truth (Manual)", "Resposta do Chat",
               "Afirmações Total", "Apoiadas", "Faithfulness", "Detalhes"]

    for caso in CASOS:
        nome = caso["nome"]
        regs = por_caso[nome]
        ws = wb.create_sheet(title=f"Caso {nome}")

        # Título da aba
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = (f"Caso {nome}  |  CHUNK_SIZE={caso['chunk_size']}  "
                   f"CHUNK_OVERLAP={caso['chunk_overlap']}")
        c.font = Font(bold=True, size=13, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Cabeçalho
        for col, h in enumerate(cab_det, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center")
            cell.border = borda_fina()

        # Dados
        for linha, r in enumerate(regs, 3):
            fila = [
                r["id"], r["pergunta"], r["ground_truth"], r["resposta_chat"],
                r["n_afirmacoes"], r["n_apoiadas"],
                f"{r['faithfulness']:.1%}", r["detalhe"]
            ]
            bg = COR_CASO.get(nome, "FFFFFF")
            for col, v in enumerate(fila, 1):
                cell = ws.cell(row=linha, column=col, value=v)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = borda_fina()
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 7:  # coluna Faithfulness
                    cell.font = Font(bold=True, color=cor_faithfulness(r["faithfulness"]))
            ws.row_dimensions[linha].height = 60

        # Larguras das colunas
        for col, w in zip(range(1, 9), [8, 35, 40, 40, 14, 12, 14, 50]):
            ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(OUTPUT_XLSX)
    print(f"\n📁 Planilha salva em: {OUTPUT_XLSX}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "="*60)
    print("   AVALIAÇÃO DO RAG — Carda TC 15")
    print("   Métrica: Faithfulness")
    print("="*60)

    # Verifica se o servidor está no ar
    print("\n[1/2] Verificando conexão com o servidor...")
    try:
        r = httpx.get("http://localhost:8000/health", timeout=5)
        data = r.json()
        print(f"  ✅ Servidor online! RAG pronto: {data.get('rag_ready')}")
    except Exception:
        print("  ❌ Servidor não está rodando!")
        print("     Abra outro terminal e execute:")
        print("     cd backend && python -m uvicorn main:app --reload")
        sys.exit(1)

    # Carrega o golden dataset
    print("\n[2/2] Carregando Golden Dataset...")
    dataset = carregar_golden_dataset()

    # Roda todos os casos
    todos_resultados = []
    for caso in CASOS:
        resultados = avaliar_caso(caso, dataset)
        todos_resultados.extend(resultados)

    # Salva a planilha
    print("\n\n💾 Salvando resultados...")
    salvar_xlsx(todos_resultados)

    # Resumo final no terminal
    from collections import defaultdict
    por_caso = defaultdict(list)
    for r in todos_resultados:
        por_caso[r["caso"]].append(r)

    print("\n" + "="*60)
    print("  RESULTADO FINAL")
    print("="*60)
    print(f"  {'Caso':<6} {'CHUNK_SIZE':<12} {'CHUNK_OVERLAP':<15} {'Faithfulness Médio'}")
    print(f"  {'-'*55}")
    melhor = max(CASOS, key=lambda c: (
        sum(r["faithfulness"] for r in por_caso[c["nome"]]) / max(len(por_caso[c["nome"]]), 1)
    ))
    for caso in CASOS:
        nome = caso["nome"]
        regs = por_caso[nome]
        media = sum(r["faithfulness"] for r in regs) / len(regs) if regs else 0
        destaque = " ◀ MELHOR" if nome == melhor["nome"] else ""
        print(f"  {nome:<6} {caso['chunk_size']:<12} {caso['chunk_overlap']:<15} {media:.1%}{destaque}")
    print("="*60)
    print(f"\n✅ Avaliação concluída! Planilha salva em:\n   {OUTPUT_XLSX}\n")


if __name__ == "__main__":
    main()
